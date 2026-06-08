#!/usr/bin/env python3
"""
Build a destination sheet for Dots scoring by copying required columns from a
source sheet and appending TotalKg and Dots formula columns.

This script resolves header names to A1 references per row, avoiding brittle
column-letter hardcoding. Provide Excel formulas with {HeaderName} placeholders
that will be replaced with the correct cell reference for each row.

Example:
  python3 scripts/sheet_builder.py \
    --workbook /root/data/openipf.xlsx \
    --source-sheet Data \
    --target-sheet Dots \
    --columns Name Sex BodyweightKg Best3SquatKg Best3BenchKg Best3DeadliftKg \
    --total-name TotalKg \
    --total-formula "ROUND({Best3SquatKg}+{Best3BenchKg}+{Best3DeadliftKg}, 3)" \
    --dots-name Dots \
    --dots-formula "ROUND({TotalKg}*500/IF(LEFT(UPPER({Sex}),1)=\"M\", <male_denom_expr>, IF(LEFT(UPPER({Sex}),1)=\"F\", <female_denom_expr>, 0)), 3)"

Notes:
- This script writes formulas but does not evaluate them. Recalculate in Excel
  or with a compatible engine after running.
- Ensure your {HeaderName} placeholders exactly match the headers written to the
  target sheet (case-sensitive).
"""

import argparse
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_args():
    p = argparse.ArgumentParser(description="Build Dots scoring sheet by header.")
    p.add_argument("--workbook", required=True, help="Path to .xlsx workbook")
    p.add_argument("--source-sheet", required=True, help="Name of source sheet (e.g., Data)")
    p.add_argument("--target-sheet", required=True, help="Name of destination sheet (e.g., Dots)")
    p.add_argument("--columns", nargs='+', required=True,
                   help="Headers to copy from source to target, in the order they appear in source")
    p.add_argument("--total-name", default="TotalKg", help="Header name for total column")
    p.add_argument("--total-formula", required=True,
                   help="Excel formula template for total column, using {Header} placeholders")
    p.add_argument("--dots-name", default="Dots", help="Header name for Dots column")
    p.add_argument("--dots-formula", required=True,
                   help="Excel formula template for Dots column, using {Header} placeholders")
    return p.parse_args()


def header_map(ws) -> Dict[str, int]:
    """Return mapping of header name -> 1-based column index from the first row."""
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val)] = col_idx
    return headers


def ensure_target_sheet(wb, name: str):
    if name in wb.sheetnames:
        # clear existing sheet by removing and recreating
        idx = wb.sheetnames.index(name)
        wb.remove(wb[name])
        ws = wb.create_sheet(title=name, index=idx)
    else:
        ws = wb.create_sheet(title=name)
    return ws


def copy_columns_by_header(src_ws, dst_ws, headers_in_order: List[str]):
    src_headers = header_map(src_ws)
    missing = [h for h in headers_in_order if h not in src_headers]
    if missing:
        raise ValueError(f"Missing required headers in source: {missing}")

    # Determine the order they appear in source to preserve relative order
    ordered = sorted(headers_in_order, key=lambda h: src_headers[h])

    # Write headers to destination
    for j, h in enumerate(ordered, start=1):
        dst_ws.cell(row=1, column=j, value=h)

    # Copy values row-wise
    for i in range(2, src_ws.max_row + 1):
        for j, h in enumerate(ordered, start=1):
            src_col = src_headers[h]
            dst_ws.cell(row=i, column=j, value=src_ws.cell(row=i, column=src_col).value)

    return ordered  # headers now in destination order


def build_ref_map(dst_ws) -> Dict[str, int]:
    """Map header name in destination to its column index (1-based)."""
    refs = {}
    for col_idx in range(1, dst_ws.max_column + 1):
        h = dst_ws.cell(row=1, column=col_idx).value
        if h is not None:
            refs[str(h)] = col_idx
    return refs


def formula_for_row(template: str, header_to_col: Dict[str, int], row: int) -> str:
    """Replace {Header} placeholders with A1 refs like C5 for the given row."""
    formula = template
    for header, col_idx in header_to_col.items():
        col_letter = get_column_letter(col_idx)
        # Replace exact placeholder occurrences
        placeholder = "{" + header + "}"
        formula = formula.replace(placeholder, f"{col_letter}{row}")
    return formula


def append_formula_column(dst_ws, header_name: str, formula_template: str):
    # Append header at next column
    next_col = dst_ws.max_column + 1
    dst_ws.cell(row=1, column=next_col, value=header_name)

    # Build mapping including the new header so subsequent columns can reference it if needed
    # But for this column's formulas, we want to reference existing columns only.
    header_to_col = {}
    for col_idx in range(1, next_col):
        h = dst_ws.cell(row=1, column=col_idx).value
        if h is not None:
            header_to_col[str(h)] = col_idx

    # Write formulas row-wise
    for row in range(2, dst_ws.max_row + 1):
        f = formula_for_row(formula_template, header_to_col, row)
        dst_ws.cell(row=row, column=next_col, value=f)


if __name__ == "__main__":
    args = parse_args()
    wb = load_workbook(args.workbook, data_only=False)

    if args.source_sheet not in wb.sheetnames:
        raise SystemExit(f"Source sheet '{args.source_sheet}' not found.")

    src_ws = wb[args.source_sheet]
    dst_ws = ensure_target_sheet(wb, args.target_sheet)

    # Copy required columns preserving source order
    copied_headers = copy_columns_by_header(src_ws, dst_ws, args.columns)

    # Append TotalKg column
    append_formula_column(dst_ws, args.total_name, args.total_formula)

    # Append Dots column
    append_formula_column(dst_ws, args.dots_name, args.dots_formula)

    wb.save(args.workbook)
    print("Sheet construction complete:")
    print(f"  Workbook: {args.workbook}")
    print(f"  Source sheet: {args.source_sheet}")
    print(f"  Target sheet: {args.target_sheet}")
    print(f"  Copied headers: {copied_headers}")
    print(f"  Added columns: {args.total_name}, {args.dots_name}")
