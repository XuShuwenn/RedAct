#!/usr/bin/env python3
"""Audit spreadsheet formula presence and cached error values.

This script is a non-destructive checker to help verify:
- That specified ranges contain formulas (INDEX/MATCH, stats, etc.)
- That cached results (if any) do not contain error tokens
- That group label cells match expected labels

It does not recalculate formulas; open the workbook in a spreadsheet app to refresh calculations if needed.

Usage:
  python3 scripts/check_sheet.py --workbook path.xlsx --sheet Task \
      --ranges C11:L20 B24:K27 C32:D41 \
      --group-label-range B9:K9 --allowed-labels Control Treated

Notes:
- Ranges are optional; provide any number of A1-style ranges.
- If the workbook has not been recalculated, cached values may be empty; the script will report such cases.
"""

import argparse
import sys
from typing import List, Tuple

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with 'pip install openpyxl'", file=sys.stderr)
    sys.exit(1)


def parse_range(a1: str) -> Tuple[str, str]:
    """Return (min_cell, max_cell) from an A1 range like 'C11:L20'."""
    parts = a1.split(":")
    if len(parts) != 2:
        raise ValueError(f"Invalid range: {a1}")
    return parts[0].strip(), parts[1].strip()


def cell_coords(cell_ref: str) -> Tuple[int, int]:
    """Convert A1 cell to (row, col)."""
    col_str = ''.join([c for c in cell_ref if c.isalpha()])
    row_str = ''.join([c for c in cell_ref if c.isdigit()])
    # Convert column letters to number
    col = 0
    for ch in col_str.upper():
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return int(row_str), col


def iter_cells(ws, a1_range: str):
    start, end = parse_range(a1_range)
    r1, c1 = cell_coords(start)
    r2, c2 = cell_coords(end)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            yield ws.cell(row=r, column=c)


def main():
    parser = argparse.ArgumentParser(description="Audit spreadsheet formulas and cached values")
    parser.add_argument("--workbook", required=True, help="Path to .xlsx workbook")
    parser.add_argument("--sheet", required=True, help="Sheet name to check")
    parser.add_argument("--ranges", nargs="*", default=[], help="A1 ranges to audit")
    parser.add_argument("--group-label-range", default=None, help="A1 range containing group labels")
    parser.add_argument("--allowed-labels", nargs="*", default=["Control", "Treated"], help="Allowed group label strings")
    args = parser.parse_args()

    # Load workbook twice: formulas view and cached values view
    try:
        wb_formulas = openpyxl.load_workbook(args.workbook, data_only=False)
        wb_values = openpyxl.load_workbook(args.workbook, data_only=True)
    except Exception as e:
        print(f"ERROR: Cannot open workbook: {e}", file=sys.stderr)
        sys.exit(1)

    if args.sheet not in wb_formulas.sheetnames:
        print(f"ERROR: Sheet '{args.sheet}' not found.", file=sys.stderr)
        sys.exit(1)

    ws_f = wb_formulas[args.sheet]
    ws_v = wb_values[args.sheet]

    # Check ranges
    total_cells = 0
    formula_cells = 0
    error_values = 0
    empty_cached = 0

    for a1 in args.ranges:
        for cf, cv in zip(iter_cells(ws_f, a1), iter_cells(ws_v, a1)):
            total_cells += 1
            # Formula detection: openpyxl stores formulas as strings starting with '='
            val_f = cf.value
            if isinstance(val_f, str) and val_f.startswith("="):
                formula_cells += 1
            # Cached values
            val_v = cv.value
            if val_v is None:
                empty_cached += 1
            elif isinstance(val_v, str) and val_v.startswith("#"):
                error_values += 1

    print("Audit Summary")
    print(f"  Sheet: {args.sheet}")
    if args.ranges:
        print(f"  Ranges checked: {', '.join(args.ranges)}")
    print(f"  Cells inspected: {total_cells}")
    print(f"  Cells with formulas: {formula_cells}")
    print(f"  Cached error tokens: {error_values}")
    print(f"  Cached empty values: {empty_cached}")

    # Group label check
    if args.group_label_range:
        labels = []
        for c in iter_cells(ws_v, args.group_label_range):
            if c.value is not None:
                labels.append(str(c.value).strip())
        if labels:
            invalid = [l for l in labels if l not in args.allowed_labels]
            print("Group Labels Check")
            print(f"  Allowed: {', '.join(args.allowed_labels)}")
            print(f"  Found:   {', '.join(labels)}")
            if invalid:
                print(f"  Invalid labels: {', '.join(invalid)}")
            else:
                print("  All labels are valid.")
        else:
            print("Group Labels Check")
            print("  No labels found in the specified range (cached values may be empty).")

    # Hints
    if error_values > 0:
        print("Hints:")
        print("  - Error tokens often indicate label mismatches (#N/A) or unsupported functions (#NAME?).")
        print("  - Verify INDEX/MATCH ranges and exact label matches; check function compatibility (STDEV.S vs STDEV).")
        print("  - Recalculate the workbook in a spreadsheet application if cached values are stale.")


if __name__ == "__main__":
    main()
