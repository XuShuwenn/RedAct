#!/usr/bin/env python3
"""
Excel Formula Audit (Validation-Only Helper)

Purpose:
  - Quickly inspect an .xlsx workbook and summarize how many cells contain formulas vs constants per sheet.
  - Helps spot sheets where calculations may have been hardcoded instead of using formulas.

Usage:
  python xlsx_formula_audit.py --file /path/to/workbook.xlsx

Notes:
  - This tool is optional and must not be used to compute or modify model values when the task requires Excel-only modeling.
  - It reads the workbook and prints a summary; it does not change the file.
"""

import argparse
import sys
from openpyxl import load_workbook


def audit_workbook(path: str) -> int:
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
    except Exception as e:
        print(f"ERROR: Failed to open workbook: {e}", file=sys.stderr)
        return 2

    print(f"Workbook: {path}")
    total_cells = 0
    total_formulas = 0

    for ws in wb.worksheets:
        formulas = 0
        constants = 0
        blanks = 0
        # Iterate a trimmed bounding box to avoid scanning huge empty grids
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is None:
                    blanks += 1
                else:
                    # In openpyxl, formulas are strings starting with '='
                    if isinstance(val, str) and val.startswith('='):
                        formulas += 1
                    else:
                        constants += 1
        sheet_total = formulas + constants + blanks
        total_cells += sheet_total
        total_formulas += formulas
        print(
            f"Sheet '{ws.title}': formulas={formulas:,} constants={constants:,} blanks={blanks:,} scanned={sheet_total:,}"
        )

    print(f"\nSummary: total_formulas={total_formulas:,} of total_scanned={total_cells:,}")
    if total_formulas == 0:
        print("WARNING: No formulas detected. Ensure calculations are not hardcoded.")
    return 0


def main():
    ap = argparse.ArgumentParser(description="Audit formulas vs constants in an Excel workbook")
    ap.add_argument("--file", required=True, help="Path to .xlsx workbook")
    args = ap.parse_args()
    sys.exit(audit_workbook(args.file))


if __name__ == "__main__":
    main()
