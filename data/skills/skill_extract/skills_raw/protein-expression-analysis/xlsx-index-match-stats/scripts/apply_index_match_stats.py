#!/usr/bin/env python3
"""
Apply two-way INDEX–MATCH formulas for a target expression grid and compute
contiguous-group statistics (means and standard deviations) and fold change.

This script is generic and parameterized via CLI arguments. It avoids
dynamic-array functions and prefers broadly compatible formulas.

Usage (dry-run to print formulas):
  python apply_index_match_stats.py --print-only \
    --workbook your.xlsx --task-sheet Task --data-sheet Data \
    --expr-top-left C11 --expr-bottom-right L20 \
    --id-col A --header-row 10 --data-id-col A --data-header-row 1 --data-last-col BZ \
    --control-cols C:G --treated-cols H:L \
    --stats-start-col B --stats-end-col K \
    --stats-row-mean-control 24 --stats-row-stdev-control 25 \
    --stats-row-mean-treated 26 --stats-row-stdev-treated 27 \
    --fc-start-row 32 --fc-end-row 41

Write formulas into workbook (omit --print-only):
  python apply_index_match_stats.py --workbook your.xlsx ...

Notes:
- Standard deviation uses STDEV(range) for compatibility. If you require
  sample vs population variants, adjust the function name accordingly.
- Fold change uses POWER(2, log2_fc) for cross-engine compatibility.
"""

import argparse
import sys
import re
from typing import Tuple

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ===== Utilities =====

def col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    idx = 0
    for c in col:
        if not ('A' <= c <= 'Z'):
            raise ValueError(f"Invalid column letter: {col}")
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx

def index_to_col_letter(index: int) -> str:
    if index <= 0:
        raise ValueError("Column index must be positive")
    result = []
    n = index
    while n > 0:
        n, r = divmod(n - 1, 26)
        result.append(chr(r + ord('A')))
    return ''.join(reversed(result))

def parse_cell_addr(addr: str) -> Tuple[int, int]:
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", addr.strip())
    if not m:
        raise ValueError(f"Invalid cell address: {addr}")
    col = col_letter_to_index(m.group(1))
    row = int(m.group(2))
    return row, col

def parse_col_range(rng: str) -> Tuple[int, int]:
    m = re.fullmatch(r"([A-Za-z]+):([A-Za-z]+)", rng.strip())
    if not m:
        raise ValueError(f"Invalid column range: {rng}")
    c1 = col_letter_to_index(m.group(1))
    c2 = col_letter_to_index(m.group(2))
    if c2 < c1:
        c1, c2 = c2, c1
    return c1, c2

# ===== Formula Builders =====

def index_match_formula(task_row: int, task_col: int, *,
                        task_header_row: int,
                        task_id_col: int,
                        data_sheet: str,
                        data_id_col_letter: str,
                        data_header_row: int,
                        data_last_col_letter: str) -> str:
    col_letter = index_to_col_letter(task_col)
    id_col_letter = index_to_col_letter(task_id_col)
    # Build references
    # INDEX(Data!$A:$BZ, MATCH($A{row}, Data!$A:$A, 0), MATCH({Col}$HeaderRow, Data!$1:$1, 0))
    rng_cols = f"$A:${data_last_col_letter}"
    data_id_col_abs = f"${data_id_col_letter}:${data_id_col_letter}"
    formula = (
        f"=IFERROR("
        f"INDEX({data_sheet}!{rng_cols},"
        f"MATCH(${id_col_letter}{task_row},{data_sheet}!{data_id_col_abs},0),"
        f"MATCH({col_letter}${task_header_row},{data_sheet}!${data_header_row}:${data_header_row},0)"
        f"),\"\")"
    )
    return formula

def mean_formula(row: int, col_start: int, col_end: int) -> str:
    a = index_to_col_letter(col_start)
    b = index_to_col_letter(col_end)
    return f"=AVERAGE({a}{row}:{b}{row})"

def stdev_formula(row: int, col_start: int, col_end: int, safe: bool = True) -> str:
    a = index_to_col_letter(col_start)
    b = index_to_col_letter(col_end)
    base = f"STDEV({a}{row}:{b}{row})"
    if not safe:
        return f"={base}"
    # Avoid #DIV/0! when fewer than 2 numeric values
    cnt = f"COUNT({a}{row}:{b}{row})"
    return f"=IF({cnt}>1,{base},\"\")"

def log2_fc_formula(stat_col_letter: str, treated_mean_row: int, control_mean_row: int) -> str:
    return f"={stat_col_letter}{treated_mean_row}-{stat_col_letter}{control_mean_row}"

def fold_change_formula(fc_row: int) -> str:
    # POWER is more portable than ^ in some contexts
    return f"=POWER(2,D{fc_row})"

# ===== Main apply routine =====

def apply_formulas(args):
    if args.print_only and not load_workbook:
        # Only printing templates; openpyxl not required
        pass

    wb = None
    ws_task = None
    if not args.print_only:
        if load_workbook is None:
            print("ERROR: openpyxl is required to write formulas.", file=sys.stderr)
            sys.exit(1)
        wb = load_workbook(args.workbook)
        ws_task = wb[args.task_sheet]

    expr_tl_row, expr_tl_col = parse_cell_addr(args.expr_top_left)
    expr_br_row, expr_br_col = parse_cell_addr(args.expr_bottom_right)
    id_col = col_letter_to_index(args.id_col)
    data_last_col_letter = args.data_last_col

    # 1) Expression block formulas
    for r in range(expr_tl_row, expr_br_row + 1):
        for c in range(expr_tl_col, expr_br_col + 1):
            f = index_match_formula(
                task_row=r, task_col=c,
                task_header_row=args.header_row,
                task_id_col=id_col,
                data_sheet=args.data_sheet,
                data_id_col_letter=args.data_id_col,
                data_header_row=args.data_header_row,
                data_last_col_letter=data_last_col_letter,
            )
            if args.print_only:
                print(f"Expr {index_to_col_letter(c)}{r}: {f}")
            else:
                ws_task.cell(row=r, column=c).value = f

    # 2) Stats formulas
    stats_start_col = col_letter_to_index(args.stats_start_col)
    stats_end_col = col_letter_to_index(args.stats_end_col)
    ctrl_start_col, ctrl_end_col = parse_col_range(args.control_cols)
    trt_start_col, trt_end_col = parse_col_range(args.treated_cols)

    n_proteins = stats_end_col - stats_start_col + 1
    for j in range(n_proteins):
        stat_col = stats_start_col + j
        # Map j-th stats column to the j-th protein row in expression block
        protein_row = expr_tl_row + j
        # Means
        f_ctrl_mean = mean_formula(protein_row, ctrl_start_col, ctrl_end_col)
        f_trt_mean = mean_formula(protein_row, trt_start_col, trt_end_col)
        # StDevs (safe by default)
        f_ctrl_sd = stdev_formula(protein_row, ctrl_start_col, ctrl_end_col, safe=True)
        f_trt_sd = stdev_formula(protein_row, trt_start_col, trt_end_col, safe=True)

        if args.print_only:
            print(f"Stats {index_to_col_letter(stat_col)}{args.stats_row_mean_control}: {f_ctrl_mean}")
            print(f"Stats {index_to_col_letter(stat_col)}{args.stats_row_stdev_control}: {f_ctrl_sd}")
            print(f"Stats {index_to_col_letter(stat_col)}{args.stats_row_mean_treated}: {f_trt_mean}")
            print(f"Stats {index_to_col_letter(stat_col)}{args.stats_row_stdev_treated}: {f_trt_sd}")
        else:
            ws_task.cell(row=args.stats_row_mean_control, column=stat_col).value = f_ctrl_mean
            ws_task.cell(row=args.stats_row_stdev_control, column=stat_col).value = f_ctrl_sd
            ws_task.cell(row=args.stats_row_mean_treated, column=stat_col).value = f_trt_mean
            ws_task.cell(row=args.stats_row_stdev_treated, column=stat_col).value = f_trt_sd

    # 3) Fold change
    if args.fc_start_row and args.fc_end_row:
        for j in range(n_proteins):
            fc_row = args.fc_start_row + j
            stat_col_letter = index_to_col_letter(stats_start_col + j)
            f_log2 = log2_fc_formula(stat_col_letter, args.stats_row_mean_treated, args.stats_row_mean_control)
            f_fc = fold_change_formula(fc_row)
            if args.print_only:
                print(f"Log2FC D{fc_row}: {f_log2}")
                print(f"FC    C{fc_row}: {f_fc}")
            else:
                # Log2 FC in column D
                ws_task.cell(row=fc_row, column=4).value = f_log2
                # FC in column C
                ws_task.cell(row=fc_row, column=3).value = f_fc

    if not args.print_only:
        wb.save(args.workbook)
        print(f"Formulas applied and workbook saved: {args.workbook}")


def main():
    p = argparse.ArgumentParser(description="Apply two-way INDEX–MATCH lookups and group stats in a spreadsheet.")
    p.add_argument('--workbook', required=False, help='Path to .xlsx workbook (required unless --print-only).')
    p.add_argument('--task-sheet', default='Task', help='Target sheet name.')
    p.add_argument('--data-sheet', default='Data', help='Data sheet name.')

    p.add_argument('--expr-top-left', required=True, help='Top-left address of expression grid (e.g., C11).')
    p.add_argument('--expr-bottom-right', required=True, help='Bottom-right address of expression grid (e.g., L20).')

    p.add_argument('--id-col', default='A', help='ID column letter on task sheet (e.g., A).')
    p.add_argument('--header-row', type=int, required=True, help='Header row index on task sheet (e.g., 10).')

    p.add_argument('--data-id-col', default='A', help='ID column letter on data sheet (e.g., A).')
    p.add_argument('--data-header-row', type=int, default=1, help='Header row index on data sheet (e.g., 1).')
    p.add_argument('--data-last-col', default='BZ', help='Last column letter to include in data sheet (e.g., BZ).')

    p.add_argument('--control-cols', required=True, help='Contiguous control column range on task sheet (e.g., C:G).')
    p.add_argument('--treated-cols', required=True, help='Contiguous treated column range on task sheet (e.g., H:L).')

    p.add_argument('--stats-start-col', required=True, help='Starting stats column letter (e.g., B).')
    p.add_argument('--stats-end-col', required=True, help='Ending stats column letter (e.g., K).')
    p.add_argument('--stats-row-mean-control', type=int, required=True, help='Row index for control means.')
    p.add_argument('--stats-row-stdev-control', type=int, required=True, help='Row index for control stdevs.')
    p.add_argument('--stats-row-mean-treated', type=int, required=True, help='Row index for treated means.')
    p.add_argument('--stats-row-stdev-treated', type=int, required=True, help='Row index for treated stdevs.')

    p.add_argument('--fc-start-row', type=int, help='Start row for fold change output (optional).')
    p.add_argument('--fc-end-row', type=int, help='End row for fold change output (optional).')

    p.add_argument('--print-only', action='store_true', help='Print formulas instead of writing to workbook.')

    args = p.parse_args()

    if not args.print_only and not args.workbook:
        print('ERROR: --workbook is required unless --print-only is set.', file=sys.stderr)
        sys.exit(1)

    apply_formulas(args)


if __name__ == '__main__':
    main()
