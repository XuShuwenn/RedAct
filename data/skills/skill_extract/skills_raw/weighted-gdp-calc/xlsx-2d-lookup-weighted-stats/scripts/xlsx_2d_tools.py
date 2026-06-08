#!/usr/bin/env python3
"""Reusable helpers to populate Excel blocks with two-key INDEX/MATCH lookups
and add ratio, stats, and weighted-mean formulas without altering formatting.

This script is generic and does NOT contain task-specific constants.
It exposes helper functions you can import into your own solution script.

Requires: openpyxl

Example usage (programmatic):

from openpyxl import load_workbook
from xlsx_2d_tools import fill_2d_index_match, add_ratio_block, add_stats, add_weighted_mean

wb = load_workbook('workbook.xlsx')
ws_task = wb['Task']

# Fill a 2D block using two-key lookup
fill_2d_index_match(
    ws=ws_task,
    top_left='H12', bottom_right='L17',
    task_row_key_col='D', task_col_key_row=9,
    data_sheet_name='Data',
    data_rowkey_col_range='B21:B40',
    data_colkey_row_range='H4:M4',
    data_values_range='H21:M40'
)

# Add ratio block (e.g., (exports-imports)/gdp * 100) to rows 35-40, columns H-L
add_ratio_block(
    ws=ws_task,
    dest_top_left='H35', dest_bottom_right='L40',
    exports_start_row=12, imports_start_row=19, gdp_start_row=26,
    multiply_by_100=True
)

# Add stats to rows 42-47 for each year column H-L based on rows 35-40
add_stats(
    ws=ws_task,
    row_min=42, row_max=43, row_median=44, row_mean=45,
    row_p25=46, row_p75=47,
    year_cols=('H','L'), data_rows=('35','40'),
    percentile_fn='PERCENTILE'
)

# Add weighted mean (e.g., GDP weights) on row 50 for H-L
add_weighted_mean(
    ws=ws_task,
    target_row=50, year_cols=('H','L'),
    data_rows_ratio=('35','40'), data_rows_weight=('26','31')
)

wb.save('workbook.xlsx')

"""

from __future__ import annotations
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
from typing import Tuple


def _addr_to_rc(addr: str) -> Tuple[int, int]:
    col, row = coordinate_from_string(addr)
    return int(row), column_index_from_string(col)


def _iter_block(top_left: str, bottom_right: str):
    r1, c1 = _addr_to_rc(top_left)
    r2, c2 = _addr_to_rc(bottom_right)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            yield r, c, get_column_letter(c)


def fill_2d_index_match(
    ws,
    top_left: str,
    bottom_right: str,
    task_row_key_col: str,
    task_col_key_row: int,
    data_sheet_name: str,
    data_rowkey_col_range: str,
    data_colkey_row_range: str,
    data_values_range: str,
):
    """Populate a destination block with INDEX/MATCH using two keys.

    - Row key: the series code in a fixed Task sheet column (e.g., 'D')
    - Column key: the year from a fixed Task sheet row (e.g., 9)

    Formula template:
    =INDEX('Data'!$VALS, MATCH($D{row}, 'Data'!$ROWKEY_COL, 0), MATCH({col}$9, 'Data'!$COLKEY_ROW, 0))
    """
    quoted_data = f"'{data_sheet_name}'" if ' ' in data_sheet_name else data_sheet_name
    for r, c, col_letter in _iter_block(top_left, bottom_right):
        row_key_ref = f"${task_row_key_col}{r}"
        col_key_ref = f"{col_letter}${task_col_key_row}"
        formula = (
            f"=INDEX({quoted_data}!${data_values_range.split(':')[0]}:${data_values_range.split(':')[1]},"
            f"MATCH({row_key_ref},{quoted_data}!${data_rowkey_col_range.split(':')[0]}:${data_rowkey_col_range.split(':')[1]},0),"
            f"MATCH({col_key_ref},{quoted_data}!${data_colkey_row_range.split(':')[0]}:${data_colkey_row_range.split(':')[1]},0))"
        )
        ws.cell(row=r, column=c).value = formula


def add_ratio_block(
    ws,
    dest_top_left: str,
    dest_bottom_right: str,
    exports_start_row: int,
    imports_start_row: int,
    gdp_start_row: int,
    multiply_by_100: bool = True,
):
    """Add per-cell ratio formulas (e.g., (exports - imports)/gdp * 100) to a block.

    Rows in destination map 1:1 to the corresponding exports/imports/gdp rows
    using the given start rows.
    """
    r1, c1 = _addr_to_rc(dest_top_left)
    r2, c2 = _addr_to_rc(dest_bottom_right)
    for i, r in enumerate(range(r1, r2 + 1)):
        exp_row = exports_start_row + i
        imp_row = imports_start_row + i
        gdp_row = gdp_start_row + i
        for c in range(c1, c2 + 1):
            col_letter = get_column_letter(c)
            formula = f"=({col_letter}{exp_row}-{col_letter}{imp_row})/{col_letter}{gdp_row}"
            if multiply_by_100:
                formula = f"{formula}*100"
            ws.cell(row=r, column=c).value = formula


def add_stats(
    ws,
    row_min: int,
    row_max: int,
    row_median: int,
    row_mean: int,
    row_p25: int,
    row_p75: int,
    year_cols: Tuple[str, str],
    data_rows: Tuple[str, str],
    percentile_fn: str = 'PERCENTILE',
):
    """Add statistics (min, max, median, mean, 25th, 75th) for each year column.

    - year_cols: (start_col_letter, end_col_letter), e.g., ('H','L')
    - data_rows: (start_row_str, end_row_str), e.g., ('35','40')
    - percentile_fn: choose 'PERCENTILE' for wider compatibility
    """
    start_col = column_index_from_string(year_cols[0])
    end_col = column_index_from_string(year_cols[1])
    rng_rows = f"{data_rows[0]}:{data_rows[1]}"
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        data_rng = f"{col_letter}{rng_rows}"
        ws.cell(row=row_min, column=c).value = f"=MIN({data_rng})"
        ws.cell(row=row_max, column=c).value = f"=MAX({data_rng})"
        ws.cell(row=row_median, column=c).value = f"=MEDIAN({data_rng})"
        ws.cell(row=row_mean, column=c).value = f"=AVERAGE({data_rng})"
        ws.cell(row=row_p25, column=c).value = f"={percentile_fn}({data_rng},0.25)"
        ws.cell(row=row_p75, column=c).value = f"={percentile_fn}({data_rng},0.75)"


def add_weighted_mean(
    ws,
    target_row: int,
    year_cols: Tuple[str, str],
    data_rows_ratio: Tuple[str, str],
    data_rows_weight: Tuple[str, str],
):
    """Add a weighted mean per year column using SUMPRODUCT.

    =SUMPRODUCT(RATIO_RANGE, WEIGHT_RANGE) / SUM(WEIGHT_RANGE)
    """
    start_col = column_index_from_string(year_cols[0])
    end_col = column_index_from_string(year_cols[1])
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        ratio_rng = f"{col_letter}{data_rows_ratio[0]}:{col_letter}{data_rows_ratio[1]}"
        weight_rng = f"{col_letter}{data_rows_weight[0]}:{col_letter}{data_rows_weight[1]}"
        ws.cell(row=target_row, column=c).value = (
            f"=SUMPRODUCT({ratio_rng},{weight_rng})/SUM({weight_rng})"
        )


if __name__ == "__main__":
    import argparse
    from openpyxl import load_workbook

    parser = argparse.ArgumentParser(description="2D lookup and stats helpers")
    parser.add_argument("workbook", help="Path to workbook (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Do not save; just open and close")
    args = parser.parse_args()

    # This CLI is intentionally minimal to discourage ad-hoc hardcoding.
    # Import and call the functions from another script where you define your ranges.
    wb = load_workbook(args.workbook)
    print("Loaded:", args.workbook, "Sheets:", wb.sheetnames)
    if not args.dry_run:
        wb.save(args.workbook)
        print("Saved (no changes by default). Import this module to use the helpers.")
