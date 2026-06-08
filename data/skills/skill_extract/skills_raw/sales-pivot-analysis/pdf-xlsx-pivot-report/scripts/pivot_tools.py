#!/usr/bin/env python3
"""Utilities to write a SourceData sheet and define standard pivot tables.

Requirements:
  - pandas
  - openpyxl >= 3.1

This module does not compute pivot values (Excel will on refresh). It defines:
  - assign_quartiles(series): robust Q1–Q4 labeling
  - write_source_and_default_pivots(df, output_path): writes SourceData and four
    pivot tables configured as:
      1) Population by State: Rows=STATE, Values=Sum(POPULATION_2023)
      2) Earners by State: Rows=STATE, Values=Sum(EARNERS)
      3) Regions by State: Rows=STATE, Values=Count(SA2_CODE)
      4) State Income Quartile: Rows=STATE, Cols=Quarter, Values=Sum(EARNERS)

Expected DataFrame columns (case-sensitive):
  ['SA2_CODE','SA2_NAME','STATE','POPULATION_2023',
   'EARNERS','MEDIAN_INCOME','MEAN_INCOME','Quarter','Total']
"""

from __future__ import annotations

from typing import Iterable, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.pivot.table import TableDefinition, Location, PivotField, DataField, RowColField
from openpyxl.pivot.cache import CacheDefinition, CacheField, CacheSource, WorksheetSource, SharedItems


REQUIRED_COLUMNS = [
    'SA2_CODE', 'SA2_NAME', 'STATE', 'POPULATION_2023',
    'EARNERS', 'MEDIAN_INCOME', 'MEAN_INCOME', 'Quarter', 'Total'
]


def assign_quartiles(series: pd.Series, labels: Iterable[str] = ("Q1", "Q2", "Q3", "Q4")) -> pd.Series:
    """Assign quartiles to a numeric series with robust fallback.

    If qcut fails (e.g., duplicate bin edges), fall back to quantile-based cut.
    Result is a string dtype with labels Q1–Q4 where possible; NaN remains NaN.
    """
    s = pd.to_numeric(series, errors="coerce")
    try:
        out = pd.qcut(s, 4, labels=list(labels))
    except Exception:
        qs = s.quantile([0.0, 0.25, 0.5, 0.75, 1.0]).dropna().unique()
        if len(qs) < 2:
            return pd.Series([pd.NA] * len(s), index=s.index, dtype="string")
        out = pd.cut(s, bins=sorted(qs), labels=list(labels)[: max(0, len(qs) - 1)], include_lowest=True, duplicates='drop')
    return out.astype("string")


def _build_cache(ref: str, sheet_name: str, columns: List[str], df: pd.DataFrame) -> CacheDefinition:
    cache_fields = []
    for col in columns:
        # Provide shared items count for likely categorical columns to improve compatibility
        if col in ("STATE", "SA2_NAME", "Quarter"):
            n = int(df[col].nunique(dropna=True)) if col in df.columns else 0
            cache_fields.append(CacheField(name=col, sharedItems=SharedItems(count=n)))
        else:
            cache_fields.append(CacheField(name=col, sharedItems=SharedItems()))
    return CacheDefinition(
        cacheSource=CacheSource(
            type="worksheet",
            worksheetSource=WorksheetSource(ref=ref, sheet=sheet_name),
        ),
        cacheFields=cache_fields,
    )


def _add_pivot(ws, cache: CacheDefinition, columns: List[str], row_fields: List[int], col_fields: List[int], data_field: int, data_name: str, data_subtotal: str, location_ref: str) -> None:
    pivot = TableDefinition(
        name=data_name.replace(" ", "").replace("-", ""),
        cacheId=0,
        dataCaption="Values",
        location=Location(ref=location_ref, firstHeaderRow=1, firstDataRow=1, firstDataCol=1),
    )
    # Build pivotFields aligned with columns order
    for idx in range(len(columns)):
        axis = None
        data_field_flag = False
        if idx in row_fields:
            axis = "axisRow"
        if idx in col_fields:
            axis = "axisCol"
        if idx == data_field:
            data_field_flag = True
        pivot.pivotFields.append(PivotField(axis=axis, dataField=data_field_flag, showAll=False))
    for idx in row_fields:
        pivot.rowFields.append(RowColField(x=idx))
    for idx in col_fields:
        pivot.colFields.append(RowColField(x=idx))
    pivot.dataFields.append(DataField(name=data_name, fld=data_field, subtotal=data_subtotal))
    pivot.cache = cache
    ws._pivots.append(pivot)


def write_source_and_default_pivots(df: pd.DataFrame, output_path: str) -> None:
    """Write SourceData and four standard pivot sheets to an Excel file.

    The input DataFrame must include REQUIRED_COLUMNS. Columns are written in the
    REQUIRED_COLUMNS order and used to index pivot fields.
    """
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Ensure types: numeric where expected, categorical as string
    for col in ["POPULATION_2023", "EARNERS", "MEDIAN_INCOME", "MEAN_INCOME"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in ["STATE", "SA2_NAME", "Quarter", "SA2_CODE"]:
        df[col] = df[col].astype("string")

    wb = Workbook()
    ws = wb.active
    ws.title = "SourceData"

    # Write header and data in REQUIRED_COLUMNS order
    ws.append(REQUIRED_COLUMNS)
    for row in df[REQUIRED_COLUMNS].itertuples(index=False):
        ws.append(list(row))

    # Determine source range
    num_rows = ws.max_row
    num_cols = ws.max_column
    ref = f"A1:{get_column_letter(num_cols)}{num_rows}"

    cache = _build_cache(ref, "SourceData", REQUIRED_COLUMNS, df)

    # Column indices for pivots
    idx_state = REQUIRED_COLUMNS.index("STATE")
    idx_pop = REQUIRED_COLUMNS.index("POPULATION_2023")
    idx_earners = REQUIRED_COLUMNS.index("EARNERS")
    idx_sa2_code = REQUIRED_COLUMNS.index("SA2_CODE")
    idx_quarter = REQUIRED_COLUMNS.index("Quarter")

    # 1) Population by State
    ws1 = wb.create_sheet("Population by State")
    _add_pivot(
        ws1, cache, REQUIRED_COLUMNS,
        row_fields=[idx_state], col_fields=[], data_field=idx_pop,
        data_name="Sum of POPULATION_2023", data_subtotal="sum", location_ref="A3:B20",
    )

    # 2) Earners by State
    ws2 = wb.create_sheet("Earners by State")
    _add_pivot(
        ws2, cache, REQUIRED_COLUMNS,
        row_fields=[idx_state], col_fields=[], data_field=idx_earners,
        data_name="Sum of EARNERS", data_subtotal="sum", location_ref="A3:B20",
    )

    # 3) Regions by State (count SA2_CODE)
    ws3 = wb.create_sheet("Regions by State")
    _add_pivot(
        ws3, cache, REQUIRED_COLUMNS,
        row_fields=[idx_state], col_fields=[], data_field=idx_sa2_code,
        data_name="Count of SA2_CODE", data_subtotal="count", location_ref="A3:B20",
    )

    # 4) State Income Quartile
    ws4 = wb.create_sheet("State Income Quartile")
    _add_pivot(
        ws4, cache, REQUIRED_COLUMNS,
        row_fields=[idx_state], col_fields=[idx_quarter], data_field=idx_earners,
        data_name="Sum of EARNERS", data_subtotal="sum", location_ref="A3:F30",
    )

    wb.save(output_path)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Write SourceData and pivot tables from a CSV/Excel file containing required columns.")
    parser.add_argument("input", help="Path to CSV or Excel file with required columns")
    parser.add_argument("output", help="Path to output Excel workbook")
    args = parser.parse_args()

    # Load input generically
    if args.input.lower().endswith(".csv"):
        df_input = pd.read_csv(args.input)
    else:
        df_input = pd.read_excel(args.input)

    # If Quarter is missing, attempt to compute from MEDIAN_INCOME
    if "Quarter" not in df_input.columns and "MEDIAN_INCOME" in df_input.columns:
        df_input["Quarter"] = assign_quartiles(df_input["MEDIAN_INCOME"]).astype("string")
    write_source_and_default_pivots(df_input, args.output)
